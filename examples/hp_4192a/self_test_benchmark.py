"""
Repeated HP 4192A self-test benchmark with Excel output.

Purpose
-------
Run the current HP 4192A self-test many times, collect reliability statistics,
and write them to one Excel workbook.

What this script does
---------------------
- runs the current `self_test` logic repeatedly
- tracks per-run pass/fail status
- tracks per-step pass/fail counts
- tracks failure categories and likely failing parameters
- stores the raw HP 4192A I/O trace for failing steps
- writes one `.xlsx` report

How to use
----------
Run from the repo root:

    python examples/hp_4192a/self_test_benchmark.py

Optional:

    python examples/hp_4192a/self_test_benchmark.py --runs 10
    python examples/hp_4192a/self_test_benchmark.py --measurement
"""

from __future__ import annotations

from collections import Counter
from contextlib import redirect_stdout
from datetime import datetime
from io import StringIO
from pathlib import Path
import argparse
import sys
import time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = REPO_ROOT / "source"
sys.path.insert(0, str(SOURCE_DIR))

import self_test as hp_self_test  # noqa: E402
from instruments import HP4192A  # noqa: E402


DEFAULT_RUNS = 50
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the HP 4192A self-test repeatedly and write an Excel report."
    )
    parser.add_argument(
        "--runs",
        type=int,
        default=DEFAULT_RUNS,
        help=f"Number of full self-test runs to execute. Default: {DEFAULT_RUNS}.",
    )
    parser.add_argument(
        "--measurement",
        action="store_true",
        help="Include the optional measurement block.",
    )
    return parser.parse_args()


def safe_sheet_name(text: str) -> str:
    invalid = set(r'[]:*?/\\')
    cleaned = "".join("_" if char in invalid else char for char in text)
    return cleaned[:31]


def capture_step_output(step_runner, *args, **kwargs) -> tuple[bool, str, Exception | None, float]:
    buffer = StringIO()
    started_at = time.monotonic()
    error: Exception | None = None

    try:
        with redirect_stdout(buffer):
            step_runner(*args, **kwargs)
    except Exception as exc:  # noqa: BLE001
        error = exc

    duration_s = time.monotonic() - started_at
    return error is None, buffer.getvalue(), error, duration_s


def append_table(worksheet, headers: list[str], rows: list[list[object]]) -> None:
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT

    for row in rows:
        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT

    adjust_column_widths(worksheet)


def adjust_column_widths(worksheet) -> None:
    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            longest_line = max((len(line) for line in value.splitlines()), default=0)
            max_length = max(max_length, longest_line)

        width = min(max(max_length + 2, 12), 80)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def build_report_path() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return DEFAULT_OUTPUT_DIR / f"self_test_benchmark_{timestamp}.xlsx"


def run_benchmark(*, runs: int, include_measurement: bool, output_path: Path) -> dict[str, object]:
    step_records: list[dict[str, object]] = []
    run_records: list[dict[str, object]] = []
    failure_details: list[dict[str, object]] = []

    step_stats: dict[tuple[str, str], dict[str, object]] = {}
    failure_categories: Counter[str] = Counter()
    failure_parameters: Counter[str] = Counter()

    all_steps: list[tuple[str, dict[str, object]]] = [("core", step) for step in hp_self_test.CORE_STEPS]
    if include_measurement:
        all_steps.extend(("measurement", step) for step in hp_self_test.MEASUREMENT_STEPS)

    benchmark_started_at = time.monotonic()

    print("HP 4192A self_test_benchmark")
    print("----------------------------")
    print(f"Resource: {hp_self_test.RESOURCE}")
    print(f"Runs: {runs}")
    print(f"Measurement block: {'enabled' if include_measurement else 'skipped'}")
    print(f"Trace enabled: {hp_self_test.ENABLE_IO_TRACE}")
    print(f"Report: {output_path}")

    for run_number in range(1, runs + 1):
        run_started_at = time.monotonic()
        run_failures = 0
        run_failed_steps: list[str] = []
        run_failure_categories: Counter[str] = Counter()
        run_failure_parameters: Counter[str] = Counter()

        print()
        print("=" * 72)
        print(f"Run {run_number} of {runs}")
        print("=" * 72)

        meter = HP4192A.open(
            hp_self_test.RESOURCE,
            timeout_ms=hp_self_test.TIMEOUT_MS,
            trace_enabled=hp_self_test.ENABLE_IO_TRACE,
            trace_print_live=hp_self_test.PRINT_LIVE_TRACE,
        )

        try:
            for step_index, (step_type, step) in enumerate(all_steps, start=1):
                step_title = str(step["title"])
                step_key = (step_type, step_title)

                if step_key not in step_stats:
                    step_stats[step_key] = {
                        "step_type": step_type,
                        "step_title": step_title,
                        "attempts": 0,
                        "passes": 0,
                        "failures": 0,
                    }

                meter.clear_trace_log()

                if step_type == "core":
                    success, output_text, error, duration_s = capture_step_output(
                        hp_self_test.run_core_step,
                        meter,
                        step,
                        step_index,
                        len(all_steps),
                    )
                else:
                    success, output_text, error, duration_s = capture_step_output(
                        hp_self_test.run_measurement_step,
                        meter,
                        step,
                        step_index,
                        len(all_steps),
                    )

                step_stats[step_key]["attempts"] += 1

                record = {
                    "run_number": run_number,
                    "step_index": step_index,
                    "step_type": step_type,
                    "step_title": step_title,
                    "status": "PASS" if success else "FAIL",
                    "duration_s": round(duration_s, 3),
                }
                step_records.append(record)

                if success:
                    step_stats[step_key]["passes"] += 1
                    print(f"[PASS] {step_type:<11} {step_title} ({duration_s:.2f} s)")
                    continue

                assert error is not None
                category = hp_self_test.classify_failure(error)
                parameter_name = hp_self_test.extract_parameter_name(error)
                trace_text = meter.format_trace_log()

                step_stats[step_key]["failures"] += 1
                failure_categories[category] += 1
                failure_parameters[parameter_name] += 1
                run_failure_categories[category] += 1
                run_failure_parameters[parameter_name] += 1
                run_failures += 1
                run_failed_steps.append(step_title)

                print(
                    f"[FAIL] {step_type:<11} {step_title} "
                    f"({duration_s:.2f} s) [{category}] {error}"
                )

                failure_details.append(
                    {
                        "run_number": run_number,
                        "step_index": step_index,
                        "step_type": step_type,
                        "step_title": step_title,
                        "category": category,
                        "parameter": parameter_name,
                        "message": str(error),
                        "configure_kwargs": repr(step.get("configure_kwargs", {})),
                        "captured_output": output_text.strip(),
                        "trace": trace_text,
                    }
                )
        finally:
            meter.close()

        run_duration_s = time.monotonic() - run_started_at
        run_status = "PASS" if run_failures == 0 else "FAIL"
        run_records.append(
            {
                "run_number": run_number,
                "status": run_status,
                "failed_steps": ", ".join(run_failed_steps),
                "failure_count": run_failures,
                "failure_categories": ", ".join(
                    f"{name}={count}" for name, count in run_failure_categories.items()
                ),
                "failure_parameters": ", ".join(
                    f"{name}={count}" for name, count in run_failure_parameters.items()
                ),
                "duration_s": round(run_duration_s, 3),
            }
        )

        print(
            f"Run result: {run_status} | failures={run_failures} | duration={run_duration_s:.2f} s"
        )

    benchmark_duration_s = time.monotonic() - benchmark_started_at

    summary = {
        "runs": runs,
        "include_measurement": include_measurement,
        "duration_s": round(benchmark_duration_s, 3),
        "run_records": run_records,
        "step_records": step_records,
        "step_stats": step_stats,
        "failure_details": failure_details,
        "failure_categories": failure_categories,
        "failure_parameters": failure_parameters,
    }

    write_excel_report(output_path=output_path, summary=summary)
    return summary


def write_excel_report(*, output_path: Path, summary: dict[str, object]) -> None:
    workbook = Workbook()

    overview_sheet = workbook.active
    overview_sheet.title = "Overview"

    run_records = summary["run_records"]
    step_stats = summary["step_stats"]
    failure_details = summary["failure_details"]
    failure_categories = summary["failure_categories"]
    failure_parameters = summary["failure_parameters"]
    total_runs = int(summary["runs"])
    passed_runs = sum(1 for record in run_records if record["status"] == "PASS")
    failed_runs = total_runs - passed_runs
    total_step_attempts = sum(stats["attempts"] for stats in step_stats.values())
    total_step_failures = sum(stats["failures"] for stats in step_stats.values())
    total_step_passes = total_step_attempts - total_step_failures

    overview_rows = [
        ["Resource", hp_self_test.RESOURCE],
        ["Runs requested", total_runs],
        ["Measurement block", "enabled" if summary["include_measurement"] else "skipped"],
        ["Trace enabled", hp_self_test.ENABLE_IO_TRACE],
        ["Trace live print", hp_self_test.PRINT_LIVE_TRACE],
        ["Benchmark duration (s)", summary["duration_s"]],
        ["Passed runs", passed_runs],
        ["Failed runs", failed_runs],
        ["Run pass rate", _format_percent(passed_runs, total_runs)],
        ["Total step attempts", total_step_attempts],
        ["Total step passes", total_step_passes],
        ["Total step failures", total_step_failures],
        ["Step pass rate", _format_percent(total_step_passes, total_step_attempts)],
        ["Failure detail rows", len(failure_details)],
    ]
    append_table(overview_sheet, ["Metric", "Value"], overview_rows)

    step_summary_sheet = workbook.create_sheet("Step Summary")
    step_summary_rows = []
    for order, stats in enumerate(step_stats.values(), start=1):
        step_summary_rows.append(
            [
                order,
                stats["step_type"],
                stats["step_title"],
                stats["attempts"],
                stats["passes"],
                stats["failures"],
                _format_percent(stats["passes"], stats["attempts"]),
            ]
        )
    append_table(
        step_summary_sheet,
        ["Order", "Step Type", "Step Title", "Attempts", "Passes", "Failures", "Pass Rate"],
        step_summary_rows,
    )

    run_summary_sheet = workbook.create_sheet("Run Summary")
    run_summary_rows = [
        [
            record["run_number"],
            record["status"],
            record["failure_count"],
            record["duration_s"],
            record["failed_steps"],
            record["failure_categories"],
            record["failure_parameters"],
        ]
        for record in run_records
    ]
    append_table(
        run_summary_sheet,
        [
            "Run",
            "Status",
            "Failure Count",
            "Duration (s)",
            "Failed Steps",
            "Failure Categories",
            "Failure Parameters",
        ],
        run_summary_rows,
    )

    category_sheet = workbook.create_sheet("Failure Categories")
    category_rows = [[name, count] for name, count in failure_categories.most_common()]
    append_table(category_sheet, ["Category", "Count"], category_rows)

    parameter_sheet = workbook.create_sheet("Failure Parameters")
    parameter_rows = [[name, count] for name, count in failure_parameters.most_common()]
    append_table(parameter_sheet, ["Parameter", "Count"], parameter_rows)

    details_sheet = workbook.create_sheet(safe_sheet_name("Failure Details"))
    detail_rows = [
        [
            item["run_number"],
            item["step_index"],
            item["step_type"],
            item["step_title"],
            item["category"],
            item["parameter"],
            item["message"],
            item["configure_kwargs"],
            item["captured_output"],
            item["trace"],
        ]
        for item in failure_details
    ]
    append_table(
        details_sheet,
        [
            "Run",
            "Step Index",
            "Step Type",
            "Step Title",
            "Category",
            "Parameter",
            "Message",
            "Configure Kwargs",
            "Captured Output",
            "Trace",
        ],
        detail_rows,
    )

    workbook.save(output_path)


def _format_percent(numerator: int, denominator: int) -> str:
    if denominator == 0:
        return "n/a"
    return f"{(100.0 * numerator / denominator):.2f}%"


def main() -> None:
    args = parse_args()
    output_path = build_report_path()

    summary = run_benchmark(
        runs=args.runs,
        include_measurement=args.measurement,
        output_path=output_path,
    )

    passed_runs = sum(1 for record in summary["run_records"] if record["status"] == "PASS")
    failed_runs = int(summary["runs"]) - passed_runs

    print()
    print("=" * 72)
    print("Benchmark summary")
    print("=" * 72)
    print(f"Passed runs: {passed_runs}")
    print(f"Failed runs: {failed_runs}")
    print(f"Report: {output_path}")


if __name__ == "__main__":
    main()
